import sys
import logging

from aiogram.fsm.state import State
from openpyxl.reader.excel import load_workbook

if sys.platform == "win32":
    import asyncio
    asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())

from bot.handlers.start_menu_handler import router
from bot.create_bot import dp, bot
from bot.handlers.authorisation import router_auth
from bot.handlers.registrartion import router_reg
from bot.handlers.auth.create_FO_po_TO import router_create
from aiogram.types import Message
from aiogram.fsm.state import StatesGroup, State
from aiogram.fsm.context import FSMContext
from aiogram.fsm.storage.memory import MemoryStorage
from openpyxl import Workbook, load_workbook
from aiogram.types import Message, ReplyKeyboardRemove, ContentType



from openpyxl import load_workbook

# def load_user_state(user_id):
#     workbook = load_workbook('states.xlsx')
#     worksheet = workbook.active
#
#     # Ищем строку с существующим состоянием пользователя
#     for row in range(2, worksheet.max_row + 1):
#         if worksheet.cell(row=row, column=1).value == user_id:
#             str_state = worksheet.cell(row=row, column=2).value
#             return str_state
#
#     # Если пользователь не найден, возвращаем None
#     return None
#
#
#
# async def main():
#     # Загружаем все роутеры
#     dp.include_routers(router, router_auth, router_reg, router_create)
#
#     # Получаем список всех пользователей из Excel
#     workbook = load_workbook('states.xlsx')
#     worksheet = workbook.active
#     for row in range(2, worksheet.max_row + 1):
#         user_id = worksheet.cell(row=row, column=1).value
#         print(user_id)
#         user_state = load_user_state(user_id)
#         if user_state:
#             # Инициализируем состояние пользователя в FSM
#             await dp.storage.set_state(chat_id=user_id, state=user_state)
#
#     # Запускаем бота
#     await dp.start_polling(bot)
#
#
# if __name__ == "__main__":
#     logging.basicConfig(level=logging.INFO)
#     try:
#         asyncio.run(main())
#     except KeyboardInterrupt:
#         print('Exit')


async def main():
    dp.include_routers(router, router_auth, router_reg, router_create)
    await dp.start_polling(bot)


if __name__ == "__main__":
    logging.basicConfig(level=logging.INFO)
    try:
        asyncio.run(main())

    except KeyboardInterrupt:
        print('Exit')

