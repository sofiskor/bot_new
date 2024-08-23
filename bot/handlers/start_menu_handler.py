import bot.markups as mp

from aiogram import Router
from aiogram.filters import CommandStart
from aiogram.types import Message
from openpyxl import load_workbook
from aiogram.fsm.state import StatesGroup, State
from aiogram.fsm.context import FSMContext



router = Router()

def save_state(user_id, state):
    # Открываем Excel-файл
    state_str = state.state
    workbook = load_workbook('states.xlsx')
    worksheet = workbook.active

    # Ищем строку с существующим состоянием пользователя
    for row in range(2, worksheet.max_row + 1):
        if worksheet.cell(row=row, column=1).value == user_id:
            # Если строка найдена, удаляем ее
            worksheet.delete_rows(row)
            break

    # Добавляем новую строку с состоянием пользователя
    worksheet.append([user_id, state_str])

    # Сохраняем Excel-файл
    workbook.save('states.xlsx')
    workbook.close()


# def get_state(user_id):
#     workbook = load_workbook('states.xlsx')
#     worksheet = workbook.active
#
#     # Ищем строку с существующим состоянием пользователя
#     for row in range(2, worksheet.max_row + 1):
#         if worksheet.cell(row=row, column=1).value == user_id:
#             # Если строка найдена, удаляем ее
#             str_state = worksheet.cell(row=row, column=2).value
#             return str_state
#     # Сохраняем Excel-файл
#     workbook.close()

class StartState(StatesGroup):
    reg_or_login = State()

# async def on_startup(_):  # информация о выходе в онлайн
#     print('Бот вышел в онлайн')


@router.message(CommandStart())
async def cmd_start(message: Message, state: FSMContext):
    await state.set_state(StartState.reg_or_login)
    save_state(message.chat.id, StartState.reg_or_login)
    await message.reply("Доброго времени суток. Для продолжения выберите пункты меню.",
                        reply_markup=mp.keyboard_start)
