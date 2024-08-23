import os

from aiogram import Router, F
from aiogram.types import Message
from openpyxl import load_workbook
from aiogram.fsm.state import StatesGroup, State
import bot.markups as mp

from bot.handlers.start_menu_handler import StartState

router_auth = Router()

def save_state(user_id, state):
    # Открываем Excel-файл
    state_str = state.state
    workbook = load_workbook('states.xlsx')
    worksheet = workbook.active

    # Ищем строку с существующим состоянием пользователя
    for row in range(2, worksheet.max_row + 1):
        if worksheet.cell(row=row, column=1).value == user_id:
            # Если строка найдена, удаляем ее
            worksheet.delete_rows(row)
            break

    # Добавляем новую строку с состоянием пользователя
    worksheet.append([user_id, state_str])

    # Сохраняем Excel-файл
    workbook.save('states.xlsx')
    workbook.close()


def get_state(user_id):
    workbook = load_workbook('states.xlsx')
    worksheet = workbook.active

    # Ищем строку с существующим состоянием пользователя
    for row in range(2, worksheet.max_row + 1):
        if worksheet.cell(row=row, column=1).value == user_id:
            # Если строка найдена, удаляем ее
            str_state = worksheet.cell(row=row, column=2).value
            print(str_state)
            state = State(str_state)
            print(state)
            return str_state
    # Сохраняем Excel-файл
    workbook.close()


# функция для проверки регистрации
def check_user_registered(user_id):
    """Функция проверяет, зарегистрирован ли пользователь по его ID."""
    if os.path.exists('registered.xlsx'):
        workbook = load_workbook('registered.xlsx')
        worksheet = workbook.active

        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if row[1] == user_id:  # ID пользователя во втором столбце
                return row[0]  # Возвращаем имя пользователя
            # print(row[1])
        workbook.close()
    return None


### Состояние Авторизация ###
@router_auth.message(F.text == "Авторизация")
async def cmd_auth(message: Message):
    user_id = message.chat.id
    previous_state_menu = get_state(user_id)
    if previous_state_menu == StartState.reg_or_login:
        user_id = message.chat.id
        user_name = check_user_registered(user_id)
        print(message.text)
        if user_name:
            await message.answer(f"welcome, {user_name}!",
                                 reply_markup=mp.auth_menu.as_markup(
                                     resize_keyboard=True)
                                 )
        else:
            await message.answer("Вы не зарегестрированы!",
                                 reply_markup=mp.keyboard_start
                                 )
    else:
        await message.answer('wrong')