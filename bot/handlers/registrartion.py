import bot.markups as mp
import os

from aiogram.fsm.state import StatesGroup, State
from aiogram import F, Router
from aiogram.fsm.context import FSMContext
from openpyxl import Workbook, load_workbook
from aiogram.types import Message, ReplyKeyboardRemove, ContentType


router_reg = Router()

# def save_state(user_id, state):
#     # Открываем Excel-файл
#     state_str = state.state
#     workbook = load_workbook('states.xlsx')
#     worksheet = workbook.active
#
#     # Ищем строку с существующим состоянием пользователя
#     for row in range(2, worksheet.max_row + 1):
#         if worksheet.cell(row=row, column=1).value == user_id:
#             # Если строка найдена, удаляем ее
#             worksheet.delete_rows(row)
#             break
#
#     # Добавляем новую строку с состоянием пользователя
#     worksheet.append([user_id, state_str])
#
#     # Сохраняем Excel-файл
#     workbook.save('states.xlsx')
#     workbook.close()
#
#
# def get_state(user_id):
#     workbook = load_workbook('states.xlsx')
#     worksheet = workbook.active
#
#     # Ищем строку с существующим состоянием пользователя
#     for row in range(2, worksheet.max_row + 1):
#         if worksheet.cell(row=row, column=1).value == user_id:
#             # Если строка найдена, удаляем ее
#             str_state = worksheet.cell(row=row, column=2).value
#             print(str_state)
#             state = State(str_state)
#             print(state)
#             return str_state
#     # Сохраняем Excel-файл
#     workbook.close()

class Registration(StatesGroup):
    waiting_for_name = State()
    waiting_for_contact = State()
    waiting_for_organization = State()


@router_reg.message(F.text == "Регистрация")
async def waiting_name(message: Message, state: FSMContext):
    await state.set_state(Registration.waiting_for_name)
    await message.reply("Введите ФИО, пожалуйста",
                        reply_markup=ReplyKeyboardRemove())


@router_reg.message(Registration.waiting_for_name)
async def waiting_name(message: Message, state: FSMContext):
    await state.update_data(user_name=message.text)
    await state.set_state(Registration.waiting_for_contact)

    await message.reply("Отправьте свой номер телефона с помощью кнопки 'Отправить контакт'.",
                        reply_markup=mp.send_contact
                        )


@router_reg.message(Registration.waiting_for_contact, F.content_type == ContentType.CONTACT)
async def waiting_contact(message: Message, state: FSMContext):
    if message.contact.user_id != message.from_user.id:
        await message.reply("Используйте кнопку для отправки контакта")
        return
    await state.update_data(user_phone=message.contact.phone_number)

    await state.set_state(Registration.waiting_for_organization)
    await message.reply("Выберите свою ПО",
                        reply_markup=mp.organisation
                        )


@router_reg.message(Registration.waiting_for_organization)
async def waiting_orgn(message: Message, state: FSMContext):
    user_orgn = message.text

    user_data = await state.get_data()
    user_name = user_data['user_name']
    user_phone = user_data['user_phone']
    user_id = message.chat.id

    if os.path.exists('registered.xlsx'):
        workbook = load_workbook('registered.xlsx')
        worksheet = workbook.active
    else:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.cell(row=1, column=1, value="Name")
        worksheet.cell(row=1, column=2, value="id")
        worksheet.cell(row=1, column=3, value="Phone")
        worksheet.cell(row=1, column=4, value="Organisation")

    row = worksheet.max_row + 1
    worksheet.cell(row=row, column=1, value=user_name)
    worksheet.cell(row=row, column=2, value=user_id)
    worksheet.cell(row=row, column=3, value=user_phone)
    worksheet.cell(row=row, column=4, value=user_orgn)

    workbook.save('registered.xlsx')  # Сохранение рабочей книги

    await message.reply(f"Ваша заявка на регистрацию "
                        f"принята и находится на проверке "
                        f"у Администратора. О результате "
                        f"проверки мы сообщим отдельно.",
                        reply_markup=mp.keyboard_start
                        )
    await state.clear()  # Завершаем состояние
