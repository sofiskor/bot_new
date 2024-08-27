import os
from aiogram.utils.keyboard import ReplyKeyboardBuilder
from aiogram.types import KeyboardButton

import bot.markups as mp

from openpyxl import load_workbook
from aiogram import Router, F
from aiogram.types import Message, ReplyKeyboardRemove
from aiogram.fsm.state import StatesGroup, State
from aiogram.fsm.context import FSMContext

from bot.handlers.start_menu_handler import cmd_start

router_create = Router()


def check_number(number):
    workbook = load_workbook('БС_Москва.xlsx')
    worksheet = workbook.active

    for row in worksheet.iter_rows(min_row=1, values_only=True):
        if row[0] == number:
            return True
    return False


def checking_number(number):
    k = 0
    if os.path.exists('data.xlsx'):
        workbook = load_workbook('data.xlsx')
        worksheet = workbook.active

        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if row[2] == number:  # ID пользователя во втором столбце
                return False
        workbook.close()
    return True


def get_data_from_data(user_id):
    workbook = load_workbook('data.xlsx')
    worksheet = workbook.active
    val = None
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if row[5] == user_id:  # ID пользователя во втором столбце
            val = (row[0], row[1], row[3], row[4])
    workbook.close()
    return val


def save_state(user_id, state, name):
    # Открываем Excel-файл
    state_str = state.state
    workbook = load_workbook('states.xlsx')
    worksheet = workbook.active

    # Ищем строку с существующим состоянием пользователя
    for row in range(2, worksheet.max_row + 1):
        if worksheet.cell(row=row, column=1).value == user_id:
            # Если строка найдена, сохраняем предыдущее состояние
            previous_state = worksheet.cell(row=row, column=2).value
            # Удаляем существующую строку
            worksheet.delete_rows(row)
            break
    else:
        # Если строка не найдена, присваиваем предыдущее состояние пустую строку
        previous_state = ''

    # Добавляем новую строку с состоянием пользователя
    worksheet.append([user_id, state_str, previous_state, name])

    # Сохраняем Excel-файл
    workbook.save('states.xlsx')
    workbook.close()


def get_state(user_id):
    workbook = load_workbook('states.xlsx')
    worksheet = workbook.active

    # Ищем строку с существующим состоянием пользователя
    for row in range(2, worksheet.max_row + 1):
        if worksheet.cell(row=row, column=1).value == user_id:
            str_state = worksheet.cell(row=row, column=2).value
            return str_state
    # Сохраняем Excel-файл
    workbook.close()


def get_name_in_state(user_id):
    workbook = load_workbook('states.xlsx')
    worksheet = workbook.active
    # Ищем строку с существующим состоянием пользователя
    for row in range(2, worksheet.max_row + 1):
        if worksheet.cell(row=row, column=1).value == user_id:
            # Если строка найдена, удаляем ее
            name_in_state = worksheet.cell(row=row, column=4).value
            return name_in_state
    # Сохраняем Excel-файл
    workbook.close()


def write_to_excel(column_number, user_id, value):
    # Загружаем существующий Excel файл
    workbook_data = load_workbook('data.xlsx')
    worksheet_data = workbook_data.active

    # Ищем существующий user_id
    last_row = worksheet_data.max_row
    existing_row = None
    for row in range(1, last_row + 1):
        if worksheet_data.cell(row=row, column=6).value == user_id:
            existing_row = row

    if existing_row:
        # Если user_id существует и 3-й столбец пустой, заполняем его
        if worksheet_data.cell(row=existing_row, column=3).value is None:
            worksheet_data.cell(row=existing_row, column=column_number, value=value)
            workbook_data.save('data.xlsx')
            workbook_data.close()
            return

    # Если user_id не найден или 3-й столбец заполнен, добавляем новую строку
    last_row = worksheet_data.max_row + 1
    worksheet_data.cell(row=last_row, column=1, value=last_row)
    worksheet_data.cell(row=last_row, column=column_number, value=value)
    worksheet_data.cell(row=last_row, column=6, value=user_id)
    print('rfgtw')
    # Сохраняем изменения в файл
    workbook_data.save('data.xlsx')
    workbook_data.close()


class CreateTO(StatesGroup):
    type_ksh = State()
    counter_number = State()
    serial_number_counter = State()
    comments = State()
    type_power = State()
    quantity_group_akb = State()
    capasity_akb = State()
    quantity_rectifier = State()
    type_rectifier = State()
    quantity_amper = State()
    quantity_system_power = State()
    check_up = State()
    exit = State()
    video = State()
    choosen_vid = State()
    ves = State()
    choosen_ksh = State()
    screenshots = State()
    akb = State()
    choosen_rrl = State()
    choosen_sector = State()
    waiting_for_video = State()
    waiting_for_number = State()
    waiting_for_quantity_antenna = State()
    waiting_for_quantity_rrl = State()
    waiting_for_choose = State()
    waiting_for_photo = State()
    choosing_sector = State()
    choosing_rrl = State()
    choosing_ksh = State()
    choosing_vid = State()


@router_create.message(F.text == "Создать ФО по ТО")
async def waiting_creation(message: Message, state: FSMContext):
    await state.set_state(CreateTO.waiting_for_number)
    save_state(message.chat.id, CreateTO.waiting_for_number, "")
    await message.reply("Введите номер базовой станции. (Например MS0001)",
                        reply_markup=ReplyKeyboardRemove())


@router_create.message(CreateTO.waiting_for_number)
async def waiting_name(message: Message, state: FSMContext):
    bs_number = message.text.upper()
    print(check_number(bs_number), checking_number(bs_number))
    if check_number(bs_number) and checking_number(bs_number):
        # await state.update_data(baseband_number=bs_number)
        user_id = message.chat.id
        user_name = None
        user_phone =None
        user_orgn = None
        # получаем данные из таблицы регистрации фио, номер и ПО
        workbook_reg = load_workbook('registered.xlsx')
        worksheet_reg = workbook_reg.active
        for row in worksheet_reg.iter_rows(min_row=2, values_only=True):
            if row[1] == user_id:  # ID пользователя во втором столбце
                user_name = row[0]
                user_phone = row[2]
                user_orgn = row[3]
        workbook_reg.close()

        # сохраняем счетчик, номер БС user_id, фио, номер и ПО в data.xlsx
        workbook_data = load_workbook('data.xlsx')
        worksheet_data = workbook_data.active
        row = worksheet_data.max_row + 1
        worksheet_data.cell(row=row, column=1, value=row - 1)
        worksheet_data.cell(row=row, column=2, value=bs_number)
        worksheet_data.cell(row=row, column=6, value=user_id)
        worksheet_data.cell(row=row, column=7, value=user_name)
        worksheet_data.cell(row=row, column=8, value=user_phone)
        worksheet_data.cell(row=row, column=9, value=user_orgn)
        workbook_data.save('data.xlsx')  # Сохранение рабочей книги
        workbook_data.close()

        await state.set_state(CreateTO.waiting_for_quantity_antenna)  # состояние ожидания кол-ва секторов
        save_state(message.chat.id, CreateTO.waiting_for_quantity_antenna, "")  # сохраняем состояние

        await message.reply("Базовой станция найдена. Выберите кол-во секторов на БС",
                            reply_markup=mp.numbers.as_markup(resize_keyboard=True)
                            )
    else:
        await message.reply("Базовой станции не найдено. Введите номер БС заново."
                            )


@router_create.message(CreateTO.waiting_for_quantity_antenna)
async def waiting_quantity_antenna(message: Message, state: FSMContext):
    if message.text.isdigit():
        bs_quantity_antenna = int(message.text)
        if bs_quantity_antenna > 20:
            await message.reply("Превышено максимальное количество секторов. "
                                "Введите новое количество секторов меньшее 20."
                                )
        else:
            # await state.update_data(bs_quantity_antenna=bs_quantity_antenna)
            write_to_excel(4, message.chat.id, bs_quantity_antenna)  # сохраняем кол-во секторов
            await state.set_state(CreateTO.waiting_for_quantity_rrl)  # состояние ожидания кол-ва РРЛ
            save_state(message.chat.id, CreateTO.waiting_for_quantity_rrl, "")  # сохраняем состояние
            await message.reply("Введите количество РРЛ на БС",
                                reply_markup=mp.numbers.as_markup(resize_keyboard=True)
                                )
    else:
        await message.answer("Пожалуйста, введите число цифрами. Например, '12'")


@router_create.message(CreateTO.waiting_for_quantity_rrl)
async def waiting_quantity_rrl(message: Message, state: FSMContext):
    if message.text.isdigit():
        bs_quantity_rrl = int(message.text)
        if bs_quantity_rrl > 25:
            await message.reply("Превышено максимальное количество РРЛ. "
                                "Введите новое количество секторов меньшее 25."
                                )
        else:
            # await state.update_data(bs_quantity_rrl=bs_quantity_rrl)
            write_to_excel(5, message.chat.id, bs_quantity_rrl)  # сохраняем кол-во РРЛ
            await state.set_state(CreateTO.waiting_for_choose)  # состояние ожидания пункта ФО по ТО
            save_state(message.chat.id, CreateTO.waiting_for_choose, "")  # сохраняем состояние
            await choose_otchet(message, state)
    else:
        await message.answer("Пожалуйста, введите число цифрами. Например, '12'")


@router_create.message(CreateTO.waiting_for_choose)
async def choose_otchet(message: Message, state: FSMContext):
    user_text = message.text
    user_id = message.chat.id

    if user_text == "Загрузить ФО по АФУ":
        await state.set_state(CreateTO.choosing_sector)
        save_state(user_id, CreateTO.choosing_sector, "")
        await waiting_sector(message, state)
    elif user_text == "Загрузить ФО по РРЛ":
        await state.set_state(CreateTO.choosing_rrl)
        save_state(user_id, CreateTO.choosing_rrl, "")
        await waiting_rrl(message, state)
    elif user_text == "Загрузить ФО по АКБ":
        await state.set_state(CreateTO.akb)
        save_state(user_id, CreateTO.akb, "")
        await waiting_akb(message, state)
    elif user_text == "Загрузить скриншоты":
        await state.set_state(CreateTO.screenshots)
        save_state(user_id, CreateTO.screenshots, "")
        await waiting_screenshots(message, state)
    elif user_text == "Загрузить ФО по КШ":
        await state.set_state(CreateTO.choosing_ksh)
        save_state(user_id, CreateTO.choosing_ksh, "")
        await waiting_ksh(message, state)
    elif user_text == "Загрузить ФО по ВЭС":
        await state.set_state(CreateTO.ves)
        save_state(user_id, CreateTO.ves, "")
        await waiting_ves(message, state)
    elif user_text == "Загрузить ФО по Общему виду":
        await state.set_state(CreateTO.choosing_vid)
        save_state(user_id, CreateTO.choosing_vid, "")
        await waiting_vid(message, state)
    elif user_text == "Загрузить видео":
        await state.set_state(CreateTO.video)
        save_state(user_id, CreateTO.video, "")
        await waiting_video(message, state)
    elif user_text == "Заполнить электронный чек-лист":
        await state.set_state(CreateTO.check_up)
        save_state(user_id, CreateTO.check_up, "")
        await waiting_check_up(message, state)
    elif user_text == "Выход без сохранения":
        await state.set_state(CreateTO.exit)
        save_state(user_id, CreateTO.exit, "")
        await waiting_exit(message, state)
    else:
        await message.answer("Пожалуйста, выберите пункт заполнения отчета с помощью кнопок",
                             reply_markup=mp.send_menu.as_markup(resize_keyboard=True)
                             )
    # await state.clear()


@router_create.message(F.text == "Назад")
async def go_back_to_otchet(message: Message, state: FSMContext):
    await state.set_state(CreateTO.waiting_for_choose)
    save_state(message.chat.id, CreateTO.waiting_for_choose, "")
    await choose_otchet(message, state)  # Возвращаемся к выбору сектора
    print('12')


@router_create.message(F.text == "Вернуться назад")
async def go_back(message: Message, state: FSMContext):
    workbook = load_workbook('states.xlsx')
    worksheet = workbook.active
    previous_state = None
    # Ищем строку с существующим состоянием пользователя
    for row in range(2, worksheet.max_row + 1):
        if worksheet.cell(row=row, column=1).value == message.chat.id:
            # Если строка найдена, удаляем ее
            previous_state = worksheet.cell(row=row, column=3).value

    # Сохраняем Excel-файл
    workbook.close()

    if previous_state == CreateTO.choosen_sector:
        await state.set_state(CreateTO.waiting_for_choose)
        save_state(message.chat.id, CreateTO.waiting_for_choose, "")
        await waiting_sector(message, state)  # Возвращаемся к выбору сектора
    elif previous_state == CreateTO.choosen_rrl:
        await state.set_state(CreateTO.waiting_for_choose)
        save_state(message.chat.id, CreateTO.waiting_for_choose, "")
        await waiting_rrl(message, state)
    elif previous_state == CreateTO.choosen_ksh:
        await state.set_state(CreateTO.waiting_for_choose)
        save_state(message.chat.id, CreateTO.waiting_for_choose, "")
        await waiting_ksh(message, state)
    elif previous_state == CreateTO.choosen_vid:
        await state.set_state(CreateTO.waiting_for_choose)
        save_state(message.chat.id, CreateTO.waiting_for_choose, "")
        await waiting_vid(message, state)
    else:
        await message.answer("Выберите вариант из кнопок")


@router_create.message(CreateTO.waiting_for_photo)
async def handle_photo(message: Message, state: FSMContext):
    if message.photo:
        photo_id = message.photo[-1].file_id  # Берем самое большое фото
        button_text = get_name_in_state(message.chat.id)
        data = get_data_from_data(message.chat.id)

        # Открываем Excel-файл
        workbook = load_workbook('Файлы.xlsx')
        worksheet = workbook.active

        # Находим первую пустую строку для записи
        next_row = worksheet.max_row + 1

        # Записываем данные в файл
        worksheet.cell(row=next_row, column=1, value=data[0])
        worksheet.cell(row=next_row, column=2, value=data[1])
        worksheet.cell(row=next_row, column=3, value=button_text)
        worksheet.cell(row=next_row, column=4, value=photo_id)

        # Сохраняем изменения
        workbook.save('Файлы.xlsx')
        workbook.close()

        await message.answer("Фото и текст кнопки успешно сохранены.")
    else:
        await message.answer("Пожалуйста, отправьте фото.")


@router_create.message(CreateTO.waiting_for_video)
async def handle_video(message: Message, state: FSMContext):
    if message.video:
        video_id = message.video.file_id  # Берем самое большое фото
        # user_data = await state.get_data()
        # button_text = user_data.get("chosen_sector")
        button_text = get_name_in_state(message.chat.id)
        data = get_data_from_data(message.chat.id)

        # Открываем Excel-файл
        workbook = load_workbook('Файлы.xlsx')
        worksheet = workbook.active

        # Находим первую пустую строку для записи
        next_row = worksheet.max_row + 1

        # Записываем данные в файл
        worksheet.cell(row=next_row, column=1, value=data[0])
        worksheet.cell(row=next_row, column=2, value=data[1])
        worksheet.cell(row=next_row, column=3, value=button_text)
        worksheet.cell(row=next_row, column=5, value=video_id)

        # Сохраняем изменения
        workbook.save('Файлы.xlsx')
        workbook.close()

        await message.answer("Видео успешно сохранено.")
    else:
        await message.answer("Пожалуйста, отправьте видео.")


## ---------- АФУ  --------------------
@router_create.message(CreateTO.choosing_sector)
async def waiting_sector(message: Message, state: FSMContext):
    datass = get_data_from_data(message.chat.id)
    builder = ReplyKeyboardBuilder()

    for i in range(1, int(datass[2]) + 1):
        builder.add(KeyboardButton(text=f"Загрузить ФО {str(i)} сектора"))
    builder.add(KeyboardButton(text="Назад"))
    builder.adjust(4)

    await message.answer(
        "Выберите по какому сектору вы хотите загрузить фотографию:",
        reply_markup=builder.as_markup(resize_keyboard=True),
    )
    # else:

    await state.set_state(CreateTO.choosen_sector)
    save_state(message.chat.id, CreateTO.choosen_sector, "")



@router_create.message(CreateTO.choosen_sector)
async def choose_sector(message: Message, state: FSMContext):
    if message.text:
        if message.text == "Вернуться назад":
            await go_back(message, state)
        else:
            name1 = message.text
            name = name1.replace("Загрузить ФО ", "")
            # await state.update_data(chosen_sector=name)  # Сохраняем выбранный сектор
            await state.set_state(CreateTO.waiting_for_photo)
            save_state(message.chat.id, CreateTO.waiting_for_photo, name)
            await message.answer("Теперь отправьте фото для выбранного сектора.",
                                 reply_markup=mp.go_back)

    else:
        await message.answer("Пожалуйста, нажмите кнопку")


## ---------- РРЛ  --------------------
@router_create.message(CreateTO.choosing_rrl)
async def waiting_rrl(message: Message, state: FSMContext):
    user_id = message.chat.id
    datass = get_data_from_data(user_id)

    builder = ReplyKeyboardBuilder()

    for i in range(1, int(datass[3]) + 1):  # количество РРЛ + 1
        builder.add(KeyboardButton(text=f"Загрузить ФО {str(i)} РРЛ"))
    builder.add(KeyboardButton(text="Назад"))
    builder.adjust(3)
    await message.answer(
        "Выберите по какой РРЛ вы хотите загрузить фотографию:",
        reply_markup=builder.as_markup(resize_keyboard=True),
    )
    await state.set_state(CreateTO.choosen_rrl)
    save_state(message.chat.id, CreateTO.choosen_rrl, "")


@router_create.message(CreateTO.choosen_rrl)
async def choose_rrl(message: Message, state: FSMContext):
    if message.text:
        if message.text == "Вернуться назад":
            await go_back(message, state)
        else:
            name1 = message.text
            name = name1.replace("Загрузить ФО ", "")
            # await state.update_data(chosen_sector=name)  # Сохраняем выбранный сектор
            await state.set_state(CreateTO.waiting_for_photo)
            save_state(message.chat.id, CreateTO.waiting_for_photo, name)
            await message.answer("Теперь отправьте фото для выбранной РРЛ.",
                                 reply_markup=mp.go_back)
    else:
        await message.answer("Пожалуйста, нажмите кнопку")


## ---------- АКБ  --------------------
@router_create.message(CreateTO.akb)
async def waiting_akb(message: Message, state: FSMContext):
    # await state.update_data(chosen_sector="АКБ")  # Сохраняем выбранный сектор
    await state.set_state(CreateTO.waiting_for_photo)
    save_state(message.chat.id, CreateTO.waiting_for_photo, "АКБ")
    await message.answer(
        "Загрузите фотографии по АКБ:",
        reply_markup=mp.go_back_menu)


## ---------- Скриншоты  --------------------
@router_create.message(CreateTO.screenshots)
async def waiting_screenshots(message: Message, state: FSMContext):
    await state.set_state(CreateTO.waiting_for_photo)
    save_state(message.chat.id, CreateTO.waiting_for_photo, "Скриншоты")
    await message.answer(
        "Загрузите cкриншоты:",
        reply_markup=mp.go_back_menu)


## ---------- КШ  --------------------
@router_create.message(CreateTO.choosing_ksh)
async def waiting_ksh(message: Message, state: FSMContext):
    await message.answer(
        "Выберите по какому пункту вы хотите загрузить фотографию:",
        reply_markup=mp.ksh_menu.as_markup(resize_keyboard=True)
    )
    await state.set_state(CreateTO.choosen_ksh)
    save_state(message.chat.id, CreateTO.choosen_ksh, "")


@router_create.message(CreateTO.choosen_ksh)
async def choose_ksh(message: Message, state: FSMContext):
    if message.text:

        if message.text == "Вернуться назад":
            await go_back(message, state)
        else:
            name1 = message.text
            name = name1.replace("Загрузить ФО ", "")
            # await state.update_data(chosen_sector=name)  # Сохраняем выбранный сектор
            await state.set_state(CreateTO.waiting_for_photo)
            save_state(message.chat.id, CreateTO.waiting_for_photo, name)
            await message.answer("Теперь отправьте фото.",
                                 reply_markup=mp.go_back)
    else:
        await message.answer("Пожалуйста, нажмите кнопку")


## ---------- ВЭС  --------------------
@router_create.message(CreateTO.ves)
async def waiting_ves(message: Message, state: FSMContext):
    await state.set_state(CreateTO.waiting_for_photo)
    save_state(message.chat.id, CreateTO.waiting_for_photo, "ВЭС")
    await message.answer(
        "Загрузите ФО по ВЭС:",
        reply_markup=mp.go_back_menu)


## ---------- Общий вид БС  --------------------
@router_create.message(CreateTO.choosing_vid)
async def waiting_vid(message: Message, state: FSMContext):
    await state.update_data(previous_state=CreateTO.choosing_vid)

    await message.answer(
        "Выберите по какому пункту вы хотите загрузить фотографии:",
        reply_markup=mp.vid_menu.as_markup(resize_keyboard=True)
    )
    await state.set_state(CreateTO.choosen_vid)
    save_state(message.chat.id, CreateTO.choosen_vid, "")


@router_create.message(CreateTO.choosen_vid)
async def choose_vid(message: Message, state: FSMContext):
    if message.text:

        if message.text == "Вернуться назад":
            await go_back(message, state)
        else:
            name1 = message.text
            name = name1.replace("Загрузить ФО ", "")
            # await state.update_data(chosen_sector=name)  # Сохраняем выбранный сектор
            await state.set_state(CreateTO.waiting_for_photo)
            save_state(message.chat.id, CreateTO.waiting_for_photo, name)
            await message.answer("Теперь отправьте фото.",
                                 reply_markup=mp.go_back)
    else:
        await message.answer("Пожалуйста, нажмите кнопку")

## ---------- Видео  --------------------
@router_create.message(CreateTO.video)
async def waiting_video(message: Message, state: FSMContext):
    # await state.update_data(chosen_sector="Видео")  # Сохраняем выбранный сектор
    await message.answer(
        "Загрузите видео",
        reply_markup=mp.go_back_menu)
    await state.set_state(CreateTO.waiting_for_video)
    save_state(message.chat.id, CreateTO.waiting_for_video, "Видео")



@router_create.message(F.text == "К предыдущему шагу")
async def previous_step(message: Message, state: FSMContext):
    workbook = load_workbook('states.xlsx')
    worksheet = workbook.active
    previous_state = None
    # Ищем строку с существующим состоянием пользователя
    for row in range(2, worksheet.max_row + 1):
        if worksheet.cell(row=row, column=1).value == message.chat.id:
            # Если строка найдена, удаляем ее
            previous_state = worksheet.cell(row=row, column=3).value
            break

    print(previous_state)
    # Сохраняем Excel-файл
    workbook.close()
    print(message.text)
    if previous_state == CreateTO.check_up:
        await state.set_state(CreateTO.check_up)
        save_state(message.chat.id, CreateTO.check_up, "")
        await waiting_check_up(message, state)  # Возвращаемся к выбору сектора
    elif previous_state == CreateTO.quantity_system_power:
        await state.set_state(CreateTO.quantity_system_power)
        save_state(message.chat.id, CreateTO.quantity_system_power, "")
        await waiting_quantity_system_power(message, state)
    elif previous_state == CreateTO.quantity_amper:
        print('2')
        await message.answer('come back')
        await state.set_state(CreateTO.quantity_amper)
        save_state(message.chat.id, CreateTO.quantity_amper, "")
        await waiting_quantity_amper(message, state)
    elif previous_state == CreateTO.quantity_rectifier:
        print('1')
        await message.answer('come back')
        await state.set_state(CreateTO.quantity_rectifier)
        save_state(message.chat.id, CreateTO.quantity_rectifier, "")
        await waiting_quantity_rectifier(message, state)
    elif previous_state == CreateTO.type_rectifier:
        await state.set_state(CreateTO.type_rectifier)
        save_state(message.chat.id, CreateTO.type_rectifier, "")
        await waiting_type_rectifier(message, state)
    # elif previous_state == CreateTO.capasity_akb:
    #     await state.set_state(CreateTO.quantity_group_akb)
    #     save_state(message.chat.id, CreateTO.quantity_group_akb, "")
    #     await waiting_quantity_group_akb(message, state)
    # elif previous_state == CreateTO.type_power:
    #     await state.set_state(CreateTO.capasity_akb)
    #     save_state(message.chat.id, CreateTO.capasity_akb, "")
    #     await waiting_capasity_akb(message, state)
    #
    # elif previous_state == CreateTO.type_ksh:
    #     await state.set_state(CreateTO.type_power)
    #     save_state(message.chat.id, CreateTO.type_power, "")
    #     await waiting_type_power(message, state)
    # elif previous_state == CreateTO.counter_number:
    #     await state.set_state(CreateTO.type_ksh)
    #     save_state(message.chat.id, CreateTO.type_ksh, "")
    #     await waiting_type_ksh(message, state)
    # elif previous_state == CreateTO.serial_number_counter:
    #     await state.set_state(CreateTO.counter_number)
    #     save_state(message.chat.id, CreateTO.counter_number, "")
    #     await waiting_counter_number(message, state)
    # elif previous_state == CreateTO.comments:
    #     await state.set_state(CreateTO.serial_number_counter)
    #     save_state(message.chat.id, CreateTO.serial_number_counter, "")
    #     await waiting_serial_number_counter(message, state)
    else:
        await message.answer("Выберите вариант из кнопок")


# заполнить чек лист
@router_create.message(CreateTO.check_up)
async def waiting_check_up(message: Message, state: FSMContext):

    print(message.text)
    await message.answer("Сколько систем питания на БС? \nОтвет введите цифрами",
                         reply_markup=ReplyKeyboardRemove())
    await state.set_state(CreateTO.quantity_system_power)
    save_state(message.chat.id, CreateTO.quantity_system_power, "")

# заполнить cистемы питания на БС
@router_create.message(CreateTO.quantity_system_power)
async def waiting_quantity_system_power(message: Message, state: FSMContext):
    print(message.text)
    num = message.text
    if message.text == "К предыдущему шагу":
        await previous_step(message, state)
    else:
        if num.isdigit() and int(num)<=3:
            write_to_excel(10, message.chat.id, int(num))
            await message.answer("Какой ток нагрузки оборудования в Амперах?",
                                 reply_markup=mp.previous_step.as_markup(resize_keyboard=True)
                                 )
            await state.set_state(CreateTO.quantity_amper)
            save_state(message.chat.id, CreateTO.quantity_amper, "")
        else:
            await message.answer("Введите ответ цифрами. Например, '2'")


# заполнить ток нагрузки в амперах
@router_create.message(CreateTO.quantity_amper)
async def waiting_quantity_amper(message: Message, state: FSMContext):
    num = message.text
    if message.text == "К предыдущему шагу":
        await previous_step(message, state)
    else:
        if num.isdigit():
            write_to_excel(11, message.chat.id, int(num))
            await message.answer("Какое количество выпрямителей?",
                                 reply_markup=mp.previous_step.as_markup(resize_keyboard=True)
                                 )
            await state.set_state(CreateTO.quantity_rectifier)
            save_state(message.chat.id, CreateTO.quantity_rectifier, "")
        else:
            await message.answer("Введите ответ цифрами. Например, '2'")


# заполнить количетсво выпрямителей
@router_create.message(CreateTO.quantity_rectifier)
async def waiting_quantity_rectifier(message: Message, state: FSMContext):
    num = message.text
    if message.text == "К предыдущему шагу":
        await previous_step(message, state)
    else:
        if num.isdigit():
            write_to_excel(12, message.chat.id, int(num))
            await message.answer("Выберите тип выпрямителя из списка:",
                                 reply_markup=mp.rectifier_menu.as_markup(resize_keyboard=True)
                                 )
            await state.set_state(CreateTO.type_rectifier)
            save_state(message.chat.id, CreateTO.type_rectifier, "")
        else:
            await message.answer("Введите ответ цифрами. Например, '2'")


# заполнить количетсво выпрямителей
@router_create.message(CreateTO.type_rectifier)
async def waiting_type_rectifier(message: Message, state: FSMContext):
    num = message.text
    l = ["Flat Pack2", "Энергомера SP-II", "Телеком-Групп", "Huawei", "Ericsson", "AEG", "Delta", "Powertel", "Штиль"]
    if message.text == "К предыдущему шагу":
        await previous_step(message, state)
    else:
        if num in l:
            write_to_excel(13, message.chat.id, num)
            await message.answer("Сколько групп АКБ на станции?",
                                 reply_markup=mp.previous_step.as_markup(resize_keyboard=True)
                                 )
            await state.set_state(CreateTO.quantity_group_akb)
            save_state(message.chat.id, CreateTO.quantity_group_akb, "")
        else:
            await message.answer("Выберите тип выпрямителя из предложенных")


# заполнить количетсво выпрямителей
@router_create.message(CreateTO.quantity_group_akb)
async def waiting_quantity_group_akb(message: Message, state: FSMContext):
    num = message.text
    if message.text == "К предыдущему шагу":
        await previous_step(message, state)
    else:
        if num.isdigit():
            write_to_excel(14, message.chat.id, int(num))
            await message.answer("Введите ёмкость АКБ числом",
                                 reply_markup=mp.previous_step.as_markup(resize_keyboard=True)
                                 )
            await state.set_state(CreateTO.capasity_akb)
            save_state(message.chat.id, CreateTO.capasity_akb, "")
        else:
            await message.answer("Введите ответ цифрами. Например, '2'")


# заполнить количетсво выпрямителей
@router_create.message(CreateTO.capasity_akb)
async def waiting_capasity_akb(message: Message, state: FSMContext):
    num = message.text
    if message.text == "К предыдущему шагу":
        await previous_step(message, state)
    else:
        if num.isdigit():
            write_to_excel(15, message.chat.id, int(num))
            await message.answer("Выберите тип источника питания",
                                 reply_markup=mp.power_type_menu.as_markup(resize_keyboard=True)
                                 )
            await state.set_state(CreateTO.type_power)
            save_state(message.chat.id, CreateTO.type_power, "")
        else:
            await message.answer("Выберите тип источника питания из предложенных")


# заполнить количетсво выпрямителей
@router_create.message(CreateTO.type_power)
async def waiting_type_power(message: Message, state: FSMContext):
    num = message.text
    if message.text:
        if message.text == "К предыдущему шагу":
            await previous_step(message, state)
        else:
            write_to_excel(16, message.chat.id, num)
            await message.answer("Выберите тип климатического шкафа из списка:",
                                 reply_markup=mp.ksh_type_menu.as_markup(resize_keyboard=True)
                                 )
            await state.set_state(CreateTO.type_ksh)
            save_state(message.chat.id, CreateTO.type_ksh, "")
    else:
        await message.answer("Пожалуйста, отправьте текст")


# заполнить количетсво выпрямителей
@router_create.message(CreateTO.type_ksh)
async def waiting_type_ksh(message: Message, state: FSMContext):
    num = message.text
    l = ["Eltek", "Ericsson", "Powertel", "Аппаратная", "Интеркросс", "Энергомера (Новая)",
         "Интеркросс Huawei", "Телеком-Групп", "Техническое помещени", "Штиль", "Энергомера"]
    if message.text == "К предыдущему шагу":
        await previous_step(message, state)
    else:
        if num in l:
            write_to_excel(17, message.chat.id, num)
            await message.answer("Введите показания электрического счётчика",
                                 reply_markup= mp.previous_step.as_markup(resize_keyboard=True)
                                 )
            await state.set_state(CreateTO.counter_number)
            save_state(message.chat.id, CreateTO.counter_number, "")
        else:
            await message.answer("Выберите тип выпрямителя из предложенных")


# заполнить количетсво выпрямителей
@router_create.message(CreateTO.counter_number)
async def waiting_counter_number(message: Message, state: FSMContext):
    num = message.text
    if message.text:
        if message.text == "К предыдущему шагу":
            await previous_step(message, state)
        else:
            write_to_excel(18, message.chat.id, num)
            await message.answer("Введите серийный номер счётчика",
                                 reply_markup=mp.previous_step.as_markup(resize_keyboard=True)
                                 )
            await state.set_state(CreateTO.serial_number_counter)
            save_state(message.chat.id, CreateTO.serial_number_counter, "")
    else:
        await message.answer("Введите ответ цифрами. Например, '2'")


# заполнить количетсво выпрямителей
@router_create.message(CreateTO.serial_number_counter)
async def waiting_serial_number_counter(message: Message, state: FSMContext):
    num = message.text
    if message.text:
        if message.text == "К предыдущему шагу":
            await previous_step(message, state)
        else:
            write_to_excel(19, message.chat.id, num)
            await message.answer("Введите замечания по недостаткам выявленные в ходе ТО. "
                                 "\nЕсли замечаний нет отправьте фразу: «Нет замечаний»",
                                 reply_markup=mp.comments.as_markup(resize_keyboard=True)
                                 )
            await state.set_state(CreateTO.comments)
            save_state(message.chat.id, CreateTO.comments, "")
    else:
        await message.answer("Введите текст")


# заполнить количетсво выпрямителей
@router_create.message(CreateTO.comments)
async def waiting_comments(message: Message, state: FSMContext):
    num = message.text
    if message.text:
        if message.text == "К предыдущему шагу":
            await previous_step(message, state)
        else:
            write_to_excel(20, message.chat.id, num)
            await message.answer("Электронный чек-лист заполнен",
                                 reply_markup=mp.comments.as_markup(resize_keyboard=True)
                                 )
            await state.set_state(CreateTO.waiting_for_choose)
            save_state(message.chat.id, CreateTO.waiting_for_choose, "")
            await waiting_sector(message, state)  # Возвращаемся к выбору сектора
    else:
        await message.answer("Введите текст")

























## ---------- Выход без сохранения  --------------------
@router_create.message(CreateTO.exit)
async def waiting_exit(message: Message, state: FSMContext):
        #      табл data
    workbook = load_workbook('data.xlsx')
    worksheet = workbook.active
    user_id =message.chat.id
    last_match_row = None
    for row in range(2, worksheet.max_row + 1):
    # Ищем последнее совпадение в 6 столбце со значением 'user_id'
        if worksheet.cell(row=row, column=6).value == user_id:
            last_match_row = row

            # Очищаем строку
    print(last_match_row)
    value_to_keep = worksheet.cell(row=last_match_row, column=1).value             # Запоминаем значение в 1 столбце

    for col in range(1, worksheet.max_column + 1):
        worksheet.cell(row=last_match_row, column=col).value = None
    # value_to_keep = worksheet.cell(row=last_match_row, column=1).value             # Запоминаем значение в 1 столбце
    workbook.save('data.xlsx')
    workbook.close()

    #      табл файлы
    workbook = load_workbook('Файлы.xlsx')
    worksheet = workbook.active
    for row in range(2, worksheet.max_row + 1):
    # Ищем последнее совпадение в 6 столбце со значением 'user_id'
        if worksheet.cell(row=row, column=1).value == value_to_keep:
            last_match_row = row
            for col in range(1, worksheet.max_column + 1):
                worksheet.cell(row=last_match_row, column=col).value = None

    # value_to_keep = worksheet.cell(row=last_match_row, column=1).value             # Запоминаем значение в 1 столбце
    workbook.save('Файлы.xlsx')
    workbook.close()
    await cmd_start(message, state)
